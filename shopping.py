from selenium import webdriver
from bs4 import BeautifulSoup
from datetime import datetime
from openpyxl import Workbook
import os
import time
from tkinter import *
from tkinter import messagebox
import threading

CHROME_LOCATION = r'C:\chromedriver_win32\chromedriver.exe'


def createFolder(directory):
    try:
        if not os.path.exists(directory):
            os.makedirs(directory)
            # print("파일생성")
    except OSError:
        print('Error: Creating directory. ' + directory)


def make_file(item_list):

    write_wb = Workbook()
    write_ws = write_wb.create_sheet('Naver Shopping')
    sheet = write_wb.active
    sheet.column_dimensions['A'].width = 100
    for i in range(len(item_list)):

        sheet.cell(i+1, 1).value = item_list[i].title
        sheet.cell(i+1, 1).hyperlink = item_list[i].url

    createFolder('/NaverShopping')
    now = time.strftime('%y%m%d-%H%M%S')
    write_wb.save(r"C:\NaverShopping\{}.xlsx".format(now))
    file_location.configure(text=r"C:\NaverShopping\{}.xlsx".format(now))


def get_category(main, sub, i):
    link = ""
    if main == 1: #해외
        if sub == 1: # 패션의류
            link = "https://search.shopping.naver.com/search/all?catId=50000000&frm=NVSHCAT&origQuery=%ED%95%B4%EC%99%B8&pagingIndex="+str(i)+"&pagingSize=40&productSet=overseas&query=%ED%95%B4%EC%99%B8&sort=review&timestamp=&viewType=list"
        elif sub == 2:
            link = "https://search.shopping.naver.com/search/all?catId=50000004&frm=NVSHCAT&origQuery=%ED%95%B4%EC%99%B8&pagingIndex="+str(i)+"&pagingSize=40&productSet=overseas&query=%ED%95%B4%EC%99%B8&sort=review&timestamp=&viewType=list"
        elif sub == 3:
            link = "https://search.shopping.naver.com/search/all?catId=50000008&frm=NVSHCAT&origQuery=%ED%95%B4%EC%99%B8&pagingIndex="+str(i)+"&pagingSize=40&productSet=overseas&query=%ED%95%B4%EC%99%B8&sort=review&timestamp=&viewType=list"
        elif sub == 4:
            link = "https://search.shopping.naver.com/search/all?catId=50000007&frm=NVSHCAT&origQuery=%ED%95%B4%EC%99%B8&pagingIndex="+str(i)+"&pagingSize=40&productSet=overseas&query=%ED%95%B4%EC%99%B8&sort=review&timestamp=&viewType=list"
        elif sub == 5:
            link = "https://search.shopping.naver.com/search/all?catId=50000006&frm=NVSHCAT&origQuery=%ED%95%B4%EC%99%B8&pagingIndex="+str(i)+"&pagingSize=40&productSet=overseas&query=%ED%95%B4%EC%99%B8&sort=review&timestamp=&viewType=list"
        elif sub == 6:
            link = "https://search.shopping.naver.com/search/all?catId=50000001&frm=NVSHCAT&origQuery=%ED%95%B4%EC%99%B8&pagingIndex="+str(i)+"&pagingSize=40&productSet=overseas&query=%ED%95%B4%EC%99%B8&sort=review&timestamp=&viewType=list"
        elif sub == 7:
            link = "https://search.shopping.naver.com/search/all?catId=50000002&frm=NVSHCAT&origQuery=%ED%95%B4%EC%99%B8&pagingIndex="+str(i)+"&pagingSize=40&productSet=overseas&query=%ED%95%B4%EC%99%B8&sort=review&timestamp=&viewType=list"
        elif sub == 8:
            link = "https://search.shopping.naver.com/search/all?catId=50000003&frm=NVSHCAT&origQuery=%ED%95%B4%EC%99%B8&pagingIndex="+str(i)+"&pagingSize=40&productSet=overseas&query=%ED%95%B4%EC%99%B8&sort=review&timestamp=&viewType=list"
        elif sub == 9:
            link = "https://search.shopping.naver.com/search/all?catId=50000005&frm=NVSHCAT&origQuery=%ED%95%B4%EC%99%B8&pagingIndex="+str(i)+"&pagingSize=40&productSet=overseas&query=%ED%95%B4%EC%99%B8&sort=review&timestamp=&viewType=list"
        elif sub == 10:
            link = "https://search.shopping.naver.com/search/all?catId=50000009&frm=NVSHCAT&origQuery=%ED%95%B4%EC%99%B8&pagingIndex="+str(i)+"&pagingSize=40&productSet=overseas&query=%ED%95%B4%EC%99%B8&sort=review&timestamp=&viewType=list"
        else:
            raise Exception("ERROR: 존재하지 않는 카테고리")
    elif main == 2: #해외직구
        if sub == 1:
            link = "https://search.shopping.naver.com/search/all?catId=50000000&frm=NVSHCAT&origQuery=%ED%95%B4%EC%99%B8%EC%A7%81%EA%B5%AC&pagingIndex="+str(i)+"&pagingSize=40&productSet=overseas&query=%ED%95%B4%EC%99%B8%EC%A7%81%EA%B5%AC&sort=review&timestamp=&viewType=list"
        elif sub == 2:
            link = "https://search.shopping.naver.com/search/all?catId=50000004&frm=NVSHCAT&origQuery=%ED%95%B4%EC%99%B8%EC%A7%81%EA%B5%AC&pagingIndex="+str(i)+"&pagingSize=40&productSet=overseas&query=%ED%95%B4%EC%99%B8%EC%A7%81%EA%B5%AC&sort=review&timestamp=&viewType=list"
        elif sub == 3:
            link = "https://search.shopping.naver.com/search/all?catId=50000008&frm=NVSHCAT&origQuery=%ED%95%B4%EC%99%B8%EC%A7%81%EA%B5%AC&pagingIndex="+str(i)+"&pagingSize=40&productSet=overseas&query=%ED%95%B4%EC%99%B8%EC%A7%81%EA%B5%AC&sort=review&timestamp=&viewType=list"
        elif sub == 4:
            link = "https://search.shopping.naver.com/search/all?catId=50000007&frm=NVSHCAT&origQuery=%ED%95%B4%EC%99%B8%EC%A7%81%EA%B5%AC&pagingIndex="+str(i)+"&pagingSize=40&productSet=overseas&query=%ED%95%B4%EC%99%B8%EC%A7%81%EA%B5%AC&sort=review&timestamp=&viewType=list"
        elif sub == 5:
            link = "https://search.shopping.naver.com/search/all?catId=50000006&frm=NVSHCAT&origQuery=%ED%95%B4%EC%99%B8%EC%A7%81%EA%B5%AC&pagingIndex="+str(i)+"&pagingSize=40&productSet=overseas&query=%ED%95%B4%EC%99%B8%EC%A7%81%EA%B5%AC&sort=review&timestamp=&viewType=list"
        elif sub == 6:
            link = "https://search.shopping.naver.com/search/all?catId=50000001&frm=NVSHCAT&origQuery=%ED%95%B4%EC%99%B8%EC%A7%81%EA%B5%AC&pagingIndex="+str(i)+"&pagingSize=40&productSet=overseas&query=%ED%95%B4%EC%99%B8%EC%A7%81%EA%B5%AC&sort=review&timestamp=&viewType=list"
        elif sub == 7:
            link = "https://search.shopping.naver.com/search/all?catId=50000002&frm=NVSHCAT&origQuery=%ED%95%B4%EC%99%B8%EC%A7%81%EA%B5%AC&pagingIndex="+str(i)+"&pagingSize=40&productSet=overseas&query=%ED%95%B4%EC%99%B8%EC%A7%81%EA%B5%AC&sort=review&timestamp=&viewType=list"
        elif sub == 8:
            link = "https://search.shopping.naver.com/search/all?catId=50000003&frm=NVSHCAT&origQuery=%ED%95%B4%EC%99%B8%EC%A7%81%EA%B5%AC&pagingIndex="+str(i)+"&pagingSize=40&productSet=overseas&query=%ED%95%B4%EC%99%B8%EC%A7%81%EA%B5%AC&sort=review&timestamp=&viewType=list"
        elif sub == 9:
            link = "https://search.shopping.naver.com/search/all?catId=50000005&frm=NVSHCAT&origQuery=%ED%95%B4%EC%99%B8%EC%A7%81%EA%B5%AC&pagingIndex="+str(i)+"&pagingSize=40&productSet=overseas&query=%ED%95%B4%EC%99%B8%EC%A7%81%EA%B5%AC&sort=review&timestamp=&viewType=list"
        elif sub == 10:
            link = "https://search.shopping.naver.com/search/all?catId=50000009&frm=NVSHCAT&origQuery=%ED%95%B4%EC%99%B8%EC%A7%81%EA%B5%AC&pagingIndex="+str(i)+"&pagingSize=40&productSet=overseas&query=%ED%95%B4%EC%99%B8%EC%A7%81%EA%B5%AC&sort=review&timestamp=&viewType=list"
        else:
            raise Exception("ERROR: 존재하지 않는 카테고리")
    elif main == 3:  # 직구
        if sub == 1:
            link = "https://search.shopping.naver.com/search/all?catId=50000000&frm=NVSHCAT&origQuery=%EC%A7%81%EA%B5%AC&pagingIndex="+str(i)+"&pagingSize=40&productSet=overseas&query=%EC%A7%81%EA%B5%AC&sort=review&timestamp=&viewType=list"
        elif sub == 2:
            link = "https://search.shopping.naver.com/search/all?catId=50000004&frm=NVSHCAT&origQuery=%EC%A7%81%EA%B5%AC&pagingIndex="+str(i)+"&pagingSize=40&productSet=overseas&query=%EC%A7%81%EA%B5%AC&sort=review&timestamp=&viewType=list"
        elif sub == 3:
            link = "https://search.shopping.naver.com/search/all?catId=50000008&frm=NVSHCAT&origQuery=%EC%A7%81%EA%B5%AC&pagingIndex="+str(i)+"&pagingSize=40&productSet=overseas&query=%EC%A7%81%EA%B5%AC&sort=review&timestamp=&viewType=list"
        elif sub == 4:
            link = "https://search.shopping.naver.com/search/all?catId=50000007&frm=NVSHCAT&origQuery=%EC%A7%81%EA%B5%AC&pagingIndex="+str(i)+"&pagingSize=40&productSet=overseas&query=%EC%A7%81%EA%B5%AC&sort=review&timestamp=&viewType=list"
        elif sub == 5:
            link = "https://search.shopping.naver.com/search/all?catId=50000006&frm=NVSHCAT&origQuery=%EC%A7%81%EA%B5%AC&pagingIndex="+str(i)+"&pagingSize=40&productSet=overseas&query=%EC%A7%81%EA%B5%AC&sort=review&timestamp=&viewType=list"
        elif sub == 6:
            link = "https://search.shopping.naver.com/search/all?catId=50000001&frm=NVSHCAT&origQuery=%EC%A7%81%EA%B5%AC&pagingIndex="+str(i)+"&pagingSize=40&productSet=overseas&query=%EC%A7%81%EA%B5%AC&sort=review&timestamp=&viewType=list"
        elif sub == 7:
            link = "https://search.shopping.naver.com/search/all?catId=50000002&frm=NVSHCAT&origQuery=%EC%A7%81%EA%B5%AC&pagingIndex="+str(i)+"&pagingSize=40&productSet=overseas&query=%EC%A7%81%EA%B5%AC&sort=review&timestamp=&viewType=list"
        elif sub == 8:
            link = "https://search.shopping.naver.com/search/all?catId=50000003&frm=NVSHCAT&origQuery=%EC%A7%81%EA%B5%AC&pagingIndex="+str(i)+"&pagingSize=40&productSet=overseas&query=%EC%A7%81%EA%B5%AC&sort=review&timestamp=&viewType=list"
        elif sub == 9:
            link = "https://search.shopping.naver.com/search/all?catId=50000005&frm=NVSHCAT&origQuery=%EC%A7%81%EA%B5%AC&pagingIndex="+str(i)+"&pagingSize=40&productSet=overseas&query=%EC%A7%81%EA%B5%AC&sort=review&timestamp=&viewType=list"
        elif sub == 10:
            link = "https://search.shopping.naver.com/search/all?catId=50000009&frm=NVSHCAT&origQuery=%EC%A7%81%EA%B5%AC&pagingIndex="+str(i)+"&pagingSize=40&productSet=overseas&query=%EC%A7%81%EA%B5%AC&sort=review&timestamp=&viewType=list"
        else:
            raise Exception("ERROR: 존재하지 않는 카테고리")
    return link


class Item:
    def __init__(self, title, url):
        self.title = title
        self.url = url

def set_today():
    return datetime.today().strftime("%Y.%m.")

def compute_date(today, date):
    date = list(map(int, date.split(".")[:-1]))
    today = list(map(int, today.split(".")[:-1]))
    month = (today[0] - date[0]) * 12 + (today[1] - date[1])
    return month

def setWay():
    global MAIN_FLAG

    way_category = way_var.get()
    if way_category == 0:
        MAIN_FLAG = False
    else:
        MAIN_FLAG = True
    print(way_category)

    return way_category


def setWhat():
    global SUB_FLAG

    what_category = what_var.get()
    print(what_category)
    if what_category == 0:
        SUB_FLAG = False
    else:
        SUB_FLAG = True

    return what_category


def setMonth():
    global MONTH_FLAG

    try:
        month = Entry.get(month_entry)
        MONTH_FLAG = True
        return int(month)
    except:
        MONTH_FLAG = False
        return


def setListCnt():
    global LIST_FLAG

    try:
        list_cnt = Entry.get(list_entry)
        LIST_FLAG = True
        return int(list_cnt)
    except:
        LIST_FLAG = False
        return None


def finalClick():
    global main_category
    global sub_category
    global MONTH
    global URL_CNT
    global MAIN_FLAG
    global SUB_FLAG
    global MONTH_FLAG
    global LIST_FLAG

    main_category = setWay()
    sub_category = setWhat()
    MONTH = setMonth()
    URL_CNT = setListCnt()
    if not(MAIN_FLAG and SUB_FLAG and MONTH_FLAG and LIST_FLAG):
        setMonth()
        messagebox.showinfo("Error", "모든 항목을 입력해주세요")
        return

    th2 = threading.Thread(target=crawling)
    th2.start()


def crawling():

    driver = webdriver.Chrome(CHROME_LOCATION)
    SCROLL_PAUSE_SEC = 0.01
    url_list = []
    title_list = []
    date_list = []
    selected_item = []
    # selected_URL_list = []
    # selected_title_list = []

    for i in range(1, 100):
        try:
            print(main_category, sub_category, i)
            link = get_category(main_category, sub_category, i)
            print(link)
            driver.get(link)

        except Exception as e:
            print(e)
            break

        last_height = driver.execute_script("return document.body.scrollHeight")
        while True:
        # 끝까지 스크롤 다운
            driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")

        # 1초 대기
            time.sleep(SCROLL_PAUSE_SEC)

        # 스크롤 다운 후 스크롤 높이 다시 가져옴
            new_height = driver.execute_script("return document.body.scrollHeight")
            if new_height == last_height:
                break
            last_height = new_height

        html = driver.page_source
        soup = BeautifulSoup(html, "html.parser")

    # 타이틀
        titles = soup.select('ul.list_basis div div div div div a[title]')

        for title in titles:
            text = title.get("title")
            title_list.append(text)
            # print(text)


        # url

        url_items = soup.select('ul.list_basis div div div div div a[href]')
        # print(str(i) + "번째")
        # print(url_items)

        tmp = None
        for url in url_items:
            text = url.get("href")

            if "https://cr" in text and text != tmp:
                url_list.append(text)
                tmp = text

        # review
        date_items = soup.select(".list_basis div span.basicList_etc__2uAYO")

        for date in date_items:
            text = date.get_text()
            # print(tmp)
            if "등록일" in text:
                text = text.split(" ")[1]
                date_list.append(text)
                # print("안쪽")
        for i, date in enumerate(date_list):
            if compute_date(today, date) <= MONTH:
                tmp = Item(title_list[i], url_list[i])
                selected_item.append(tmp)

        if len(selected_item) > URL_CNT:  # 가져오려는 리스트수
            driver.close()
            selected_item = selected_item[:URL_CNT]
            break
    for i in range(URL_CNT):
        print(selected_item)
        print(str(i + 1) + ". " + selected_item[i].title +"\n URL: ", selected_item[i].url + "\n")

    print()

    make_file(selected_item)

# 사용하지 않는 기능
    # print("< 상품 필터링하기 >")
    # while(True):
    #     tmp_list = []
    #     x = input("빼고 싶은 상품 입력(-1입력시 종료): ")
    #     y = 0
    #     if x == "-1":
    #         print("종료합니다")
    #
    #         break
    #
    #     i = 0
    #     for i in range(len(selected_item)):
    #         if(x not in selected_item[i].title):
    #             tmp_list.append(selected_item[i])

    #     if y == len(selected_item):
    #         print("상품 없음")
    #         break
    #     selected_item = tmp_list
    #
    # print()
    # print("< 필터링된 상품 목록 >")
    # for i in range(len(selected_item)):
    #     print(str(i + 1) + ". " + selected_item[i].title +"\n URL: ", selected_item[i].url + "\n")




MAIN_FLAG = False
SUB_FLAG = False
MONTH_FLAG = False
LIST_FLAG = False

today = set_today()
print(set_today())

global main_category
global sub_category
global MONTH
global URL_CNT
WAY = ["해외", "해외직구", "직구"]
WHAT = ["패션의류", "가구/인테리어", "생활/건강", "스포츠/레저", "식품", "패션잡화", "화장품/미용", "디지털/가전", "출산/육아", "여가/생활편의"]
window = Tk()
window.geometry("360x440")
window.title("Naver Shopping Crawling")
window.resizable(width=False, height=False)
way_btn = []
what_btn = []
way_var = IntVar()


# 메인 카테고리
for i, way_title in enumerate(WAY):
    way_btn.append(Radiobutton(window, text=way_title, variable=way_var, value=i + 1, command=setWay))
x = 20
y = 10
main_label = Label(window, text= "< Main Category >")
main_label.place(x=x, y=y)
y += 25
for i in range(len(WAY)):
    way_btn[i].place(x=x, y=y)
    x += 100
what_var = IntVar()
y += 50
for i, what_title in enumerate(WHAT):
    what_btn.append(Radiobutton(window, text=what_title, variable=what_var, value=i + 1, command=setWhat))
x = 20


# 서브 카테고리
sub_label = Label(window, text="< Sub Category >")
sub_label.place(x=x, y=y)
y += 25
tmp = y
for i in range(len(WHAT)//2+1):
    what_btn[i].place(x=x, y=y)
    y += 30
x = 220
y = tmp
for i in range(len(WHAT)//2, len(WHAT)):
    what_btn[i].place(x=x, y=y)
    y += 30


# 몇개월이내로 가져올지 결정하기
x = 25
y = 270
month_text = Label(window, text="몇개월이내로 가져오시겠습니까?")
month_text.place(x=x, y=y)
y += 25
month_entry = Entry(window, width=5)
month_entry.place(x=x+10, y=y)
mon_label = Label(window, text="개월")
mon_label.place(x=x+50, y=y)
y += 40


# 몇개의 리스트 가져올지 결정하기
list_cnt = Label(window, text="몇개의 리스트를 가져오시겠습니까?")
list_entry = Entry(window, width=5)
list_cnt.place(x=x, y=y)
y += 25
list_entry.place(x=x+10, y=y)
mon_label = Label(window, text="개")
mon_label.place(x=x+50, y=y)


# 크롤링하기 버튼
label = Button(window, text="크롤링", width=10, height=5,  command=finalClick)
label.place(x=250, y=280)


# 만들어진 엑셀 파일 주소
file_location = Label(window, text="파일주소:")
file_location.place(x=30, y=400)

window.mainloop()


















